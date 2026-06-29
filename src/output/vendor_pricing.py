"""
src/output/vendor_pricing.py
=============================
Phase 6 — Vendor Pricing Ingestion & Proposal Generation

Steps:
  6.1  ingest_vendor_pricing()  — read a filled vendor inquiry CSV/XLSX and
                                   match prices back to BOM rows
  6.2  finalize_bom_pricing()   — recalculate Total Price and set DATA_COMPLETE_FLAG
  6.3  write_priced_annexure()  — export the final priced BOM CSV (Annexure-I)
  6.4  write_revision()         — re-export after customer change requests
"""

from __future__ import annotations

import csv
import logging
import os
import re
from typing import Dict, List

import openpyxl

from src.core.models import RMCData
from src.output.csv_writer import write_annexure_csv, write_spares_summary_csv

log = logging.getLogger(__name__)


# ── Phase 6.1: Vendor price ingestion ────────────────────────────────────────

def _read_vendor_csv(path: str) -> List[Dict[str, str]]:
    """Read a vendor-filled CSV and return list of row dicts."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({k.strip().upper(): v.strip() for k, v in row.items()})
    return rows


def _read_vendor_xlsx(path: str) -> List[Dict[str, str]]:
    """Read a vendor-filled XLSX — assumes row 3 is the header (matching our template)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().upper() if c.value else "" for c in ws[3]]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(row):
            continue
        rows.append({
            headers[i]: (str(v).strip() if v is not None else "")
            for i, v in enumerate(row)
            if i < len(headers) and headers[i]
        })
    return rows


def ingest_vendor_pricing(vendor_file: str) -> List[Dict[str, str]]:
    """
    Phase 6.1: Read the vendor-filled inquiry file (CSV or XLSX).
    Returns a list of dicts with at minimum:
      INTERNAL ITEM CODE, MAKE, UNIT PRICE, DELIVERY TIME (WEEKS),
      OFFER REF, VALIDITY DATE
    """
    ext = os.path.splitext(vendor_file)[1].lower()
    if ext in (".xlsx", ".xls"):
        rows = _read_vendor_xlsx(vendor_file)
    elif ext == ".csv":
        rows = _read_vendor_csv(vendor_file)
    else:
        raise ValueError(f"Unsupported vendor file format: {ext}")

    log.info("Ingested %d vendor pricing rows from %s", len(rows), vendor_file)
    return rows


# ── Phase 6.1 cont.: match prices to BOM rows ─────────────────────────────────

def _normalise(s: str) -> str:
    """Lower-case, remove punctuation, collapse whitespace for fuzzy matching."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", "", s.lower())).strip()


def apply_vendor_pricing(bom_rows: List[dict], vendor_rows: List[dict]) -> List[dict]:
    """
    Phase 6.1: Match each vendor row to a BOM row and populate:
      make, unit_price, delivery_weeks, offer_ref, validity_date.
    Matching strategy (in order):
      1. INTERNAL ITEM CODE exact match
      2. Normalised description token overlap (≥ 3 common tokens)

    Returns the updated bom_rows list.
    """
    # Build a lookup by item code from vendor data
    code_lookup: Dict[str, dict] = {}
    for vr in vendor_rows:
        code = vr.get("INTERNAL ITEM CODE", "").strip()
        if code:
            code_lookup[code.upper()] = vr

    for row in bom_rows:
        code = (row.get("internal_item_code") or "").upper()
        matched = code_lookup.get(code)

        # Fallback: fuzzy description match
        if not matched:
            desc_tokens = set(_normalise(row.get("item_description", "")).split())
            best_overlap = 0
            for vr in vendor_rows:
                v_desc_tokens = set(
                    _normalise(
                        vr.get("PART DESCRIPTION", "") or
                        vr.get("ITEM (DESCRIPTION)", "") or ""
                    ).split()
                )
                overlap = len(desc_tokens & v_desc_tokens)
                if overlap > best_overlap and overlap >= 3:
                    best_overlap = overlap
                    matched = vr

        if matched:
            row["make"] = matched.get("MAKE", "")
            row["unit_price"] = matched.get("UNIT PRICE", "")
            row["delivery_weeks"] = matched.get("DELIVERY TIME (WEEKS)", "")
            row["offer_ref"] = matched.get("OFFER REF", "")
            row["validity_date"] = matched.get("VALIDITY DATE", "")

    log.info("Vendor pricing applied to %d BOM rows.", len(bom_rows))
    return bom_rows


# ── Phase 6.2: Finalize pricing ───────────────────────────────────────────────

def finalize_bom_pricing(bom_rows: List[dict]) -> List[dict]:
    """
    Phase 6.2: For each row with a unit price, calculate Total Price = Qty × Unit Price.
    Set DATA_COMPLETE_FLAG = True once make + price are filled.
    """
    for row in bom_rows:
        qty_str = row.get("qty", "").strip()
        price_str = row.get("unit_price", "").strip()

        if qty_str and price_str:
            try:
                qty = float(re.sub(r"[^\d.]", "", qty_str))
                price = float(re.sub(r"[^\d.]", "", price_str))
                row["total_price"] = f"{qty * price:,.2f}"
            except ValueError:
                row["total_price"] = ""
        else:
            row["total_price"] = ""

        # Update data completeness flag
        required = ["item_description", "qty", "make", "unit_price", "delivery_weeks"]
        row["data_complete_flag"] = all(row.get(f, "").strip() for f in required)

    return bom_rows


# ── Phase 6.3: Export priced BOM ──────────────────────────────────────────────

def write_priced_annexure(
    rmc: RMCData,
    bom_rows: List[dict],
    bom_root,
    warnings: List[str],
    out_dir: str,
    lead_no: str,
    rfq_ref: str,
    project: str,
    customer: str,
    date_rcv: str,
    meta=None,
) -> dict:
    """
    Phase 6.3: Export the final priced BOM CSV and customer spares summary.
    Returns a dict with the output file paths.
    """
    def _slug(s: str) -> str:
        return re.sub(r"[^A-Za-z0-9]+", "", s.upper())[:12]

    bom_fname = "_".join(filter(None, [
        "PRICED_BOM", lead_no, _slug(rfq_ref), _slug(project),
        _slug(customer), re.sub(r"[^A-Za-z0-9]", "", date_rcv)[:9],
    ])) + ".csv"

    spares_fname = "_".join(filter(None, [
        "SPARES_SUMMARY", lead_no, _slug(rfq_ref),
    ])) + ".csv"

    bom_path = os.path.join(out_dir, bom_fname)
    spares_path = os.path.join(out_dir, spares_fname)

    write_annexure_csv(rmc, bom_rows, bom_root, warnings, bom_path, meta)
    write_spares_summary_csv(bom_rows, rmc, spares_path, meta)

    log.info("Priced BOM CSV: %s", bom_path)
    log.info("Spares summary CSV: %s", spares_path)
    return {"bom_csv": bom_path, "spares_csv": spares_path}


# ── Phase 6.4: Revision handling ─────────────────────────────────────────────

def write_revision(
    rmc: RMCData,
    bom_rows: List[dict],
    bom_root,
    warnings: List[str],
    out_dir: str,
    lead_no: str,
    rfq_ref: str,
    project: str,
    customer: str,
    date_rcv: str,
    revision: int = 1,
    meta=None,
) -> str:
    """
    Phase 6.4: Write a revised BOM CSV (Rev 1, Rev 2 …) after customer
    change requests, keeping the original file intact.
    Returns the path to the new file.
    """
    def _slug(s: str) -> str:
        return re.sub(r"[^A-Za-z0-9]+", "", s.upper())[:12]

    fname = "_".join(filter(None, [
        "PRICED_BOM", lead_no, _slug(rfq_ref), _slug(project),
        _slug(customer), f"Rev{revision}",
    ])) + ".csv"

    path = os.path.join(out_dir, fname)
    write_annexure_csv(rmc, bom_rows, bom_root, warnings, path, meta)
    log.info("Revision %d BOM CSV written: %s", revision, path)
    return path
