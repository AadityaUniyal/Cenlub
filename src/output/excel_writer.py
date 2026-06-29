"""
src/output/excel_writer.py
===========================
Generates a styled three-tab Excel workbook from RMCData:
  Tab 1 – Master RMC Parameters
  Tab 2 – Mandatory Spares
  Tab 3 – O&M Spares
"""

from __future__ import annotations

import os
from typing import List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.models import RMCData


# ── Style helpers ─────────────────────────────────────────────────────────────

def _title_banner(ws, text: str, colour: str = "1F4E78") -> None:
    ws.merge_cells("A1:D2")
    cell = ws["A1"]
    cell.value = text
    cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20


def _header_row(ws, row: int, headers: List[str], colour: str) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")


_THIN_BORDER = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)


def _apply_style(ws) -> None:
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                if not cell.font or not cell.font.bold:
                    cell.font = Font(name="Calibri", size=10)
                cell.border = _THIN_BORDER


def _autofit(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.row > 2 and cell.value),
            default=0,
        )
        ws.column_dimensions[letter].width = max(max_len + 4, 12)


# ── Public API ────────────────────────────────────────────────────────────────

def write_excel(data: RMCData, out_path: str) -> None:
    """Write a styled Excel workbook to *out_path*."""
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb = openpyxl.Workbook()

    # ── Tab 1: Master RMC ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Master RMC"
    _title_banner(ws1, f"Requirement Mapping Checklist (RMC) — Lead: {data.lead_no}")
    _header_row(ws1, 4, ["Engineering Parameter", "Extracted Value"], "2F5597")

    rows: List[Tuple[str, str]] = [
        ("Lead Number", data.lead_no),
        ("Project Name", data.project_name),
        ("Customer Name", data.customer),
        ("Application", data.application),
        ("Industry", data.industry),
        ("Flow Rate", data.flow_rate),
        ("Operating Pressure", data.pressure),
        ("Operating Temperature", data.temperature),
        ("Oil Grade", data.oil_grade),
        ("Filtration Level", data.filtration_level),
        ("Heat Exchanger MOC", data.heat_exchanger_moc),
        ("Reservoir MOC", data.reservoir_moc),
        ("Control Panel Requirement", data.control_panel_requirement),
    ]
    for idx, (param, val) in enumerate(rows):
        r = 5 + idx
        ws1.cell(row=r, column=1, value=param)
        vc = ws1.cell(row=r, column=2, value=val)
        if val == "XX":
            vc.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            vc.font = Font(color="C00000", name="Calibri", size=10)

    _apply_style(ws1)
    _autofit(ws1)

    # ── Tab 2: Mandatory Spares ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Mandatory Spares")
    _title_banner(ws2, "Mandatory Spares Checklist", "C53030")
    _header_row(ws2, 4, ["Item No", "Description", "Quantity", "Part Number"], "9B2C2C")
    for i, s in enumerate(data.mandatory_spares):
        r = 5 + i
        ws2.cell(row=r, column=1, value=s.item_no)
        ws2.cell(row=r, column=2, value=s.description)
        ws2.cell(row=r, column=3, value=s.qty)
        ws2.cell(row=r, column=4, value=s.part_no)
    _apply_style(ws2)
    _autofit(ws2)

    # ── Tab 3: O&M Spares ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("O&M Spares")
    _title_banner(ws3, "Operation & Maintenance (O&M) Spares", "D69E2E")
    _header_row(ws3, 4, ["Item No", "Description", "Quantity", "Frequency"], "B7791F")
    for i, s in enumerate(data.om_spares):
        r = 5 + i
        ws3.cell(row=r, column=1, value=s.item_no)
        ws3.cell(row=r, column=2, value=s.description)
        ws3.cell(row=r, column=3, value=s.qty)
        ws3.cell(row=r, column=4, value=s.frequency)
    _apply_style(ws3)
    _autofit(ws3)

    wb.save(out_path)
