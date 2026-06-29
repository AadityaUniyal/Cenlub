"""
src/output/vendor_inquiry.py
=============================
Phase 3.4 — Vendor Inquiry Package Generation

Generates:
  1. vendor_inquiry_<lead>_<rfq>.csv
     One row per applicable BOM item, with Item Code, Description, Qty,
     Technical Spec, MOC. Make/Price/Delivery left blank for vendor fill.
  2. vendor_inquiry_<lead>_<rfq>.xlsx
     Styled Excel version of the above (easier for vendors to fill in).

Phase 3.5 — Auto-Reminder Emails
  send_vendor_inquiry()  : sends the inquiry to a list of vendor emails via SMTP
  send_vendor_reminder() : re-sends a reminder if no response after SLA days
"""

from __future__ import annotations

import csv
import logging
import os
import smtplib
from datetime import date, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Column layout for vendor inquiry CSV ─────────────────────────────────────

VENDOR_HEADERS = [
    "SR.NO",
    "INTERNAL ITEM CODE",
    "CATEGORY",
    "PART DESCRIPTION",
    "PLAIN DESCRIPTION",
    "QTY",
    "UNIT",
    "TECHNICAL SPECIFICATION",
    "MOC",
    "MAKE",           # vendor fills
    "UNIT PRICE",     # vendor fills
    "TOTAL PRICE",    # vendor fills (or auto-calc)
    "DELIVERY TIME (WEEKS)",  # vendor fills
    "OFFER REF",              # vendor fills
    "VALIDITY DATE",          # vendor fills
    "REMARKS",
]


# ── CSV writer ────────────────────────────────────────────────────────────────

def write_vendor_inquiry_csv(
    bom_rows: List[dict],
    lead_no: str,
    rfq_ref: str,
    project: str,
    customer: str,
    out_path: str,
) -> None:
    """
    Phase 3.4: Write the vendor inquiry CSV — applicable items only,
    price/make/delivery columns blank for vendor completion.
    """
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)

    applicable = [r for r in bom_rows if r.get("applicable_flag", True)]

    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        # Mini header
        w.writerow(["VENDOR INQUIRY — CENLUB SYSTEMS"])
        w.writerow(["Lead No", lead_no])
        w.writerow(["RFQ Ref", rfq_ref])
        w.writerow(["Project", project])
        w.writerow(["Customer", customer])
        w.writerow(["Date", date.today().strftime("%d-%b-%Y")])
        w.writerow([])
        w.writerow(VENDOR_HEADERS)

        for i, row in enumerate(applicable, 1):
            cat = f"{row.get('category_code', '')} – {row.get('category_label', '')}".strip(" –")
            w.writerow([
                i,
                row.get("internal_item_code", ""),
                cat,
                row.get("item_description", ""),
                row.get("plain_description", ""),
                row.get("qty", ""),
                row.get("unit", ""),
                row.get("technical_specification", ""),
                row.get("moc", ""),
                "",  # MAKE — vendor fills
                "",  # UNIT PRICE — vendor fills
                "",  # TOTAL PRICE — vendor fills
                "",  # DELIVERY — vendor fills
                "",  # OFFER REF — vendor fills
                "",  # VALIDITY DATE — vendor fills
                row.get("remarks", ""),
            ])

    log.info("Vendor inquiry CSV written: %s (%d applicable items)", out_path, len(applicable))


# ── Excel writer ──────────────────────────────────────────────────────────────

_THIN = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)


def write_vendor_inquiry_xlsx(
    bom_rows: List[dict],
    lead_no: str,
    rfq_ref: str,
    project: str,
    customer: str,
    out_path: str,
) -> None:
    """
    Phase 3.4: Write the vendor inquiry as a styled Excel file.
    Blank price/make/delivery cells are highlighted yellow for easy identification.
    """
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendor Inquiry"

    applicable = [r for r in bom_rows if r.get("applicable_flag", True)]

    # Title block
    ws.merge_cells("A1:P2")
    title_cell = ws["A1"]
    title_cell.value = (
        f"VENDOR INQUIRY — CENLUB SYSTEMS  |  Lead: {lead_no}"
        f"  |  RFQ: {rfq_ref}  |  {project} / {customer}"
    )
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18

    # Header row
    hdr_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    for col, h in enumerate(VENDOR_HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    vendor_cols = {10, 11, 12, 13, 14, 15}   # MAKE, UNIT PRICE, TOTAL PRICE, DELIVERY, OFFER, VALIDITY

    for i, row in enumerate(applicable, 1):
        r = 3 + i
        cat = f"{row.get('category_code', '')} – {row.get('category_label', '')}".strip(" –")
        values = [
            i,
            row.get("internal_item_code", ""),
            cat,
            row.get("item_description", ""),
            row.get("plain_description", ""),
            row.get("qty", ""),
            row.get("unit", ""),
            row.get("technical_specification", ""),
            row.get("moc", ""),
            "",  # MAKE
            "",  # UNIT PRICE
            "",  # TOTAL PRICE
            "",  # DELIVERY
            "",  # OFFER REF
            "",  # VALIDITY DATE
            row.get("remarks", ""),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Calibri", size=9)
            c.border = _THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col in vendor_cols:
                c.fill = yellow

    # Auto-fit columns
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.row > 2 and cell.value),
            default=0,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 4, 10), 45)

    ws.freeze_panes = "A4"
    wb.save(out_path)
    log.info("Vendor inquiry XLSX written: %s", out_path)


# ── Email sender (Phase 3.4) ──────────────────────────────────────────────────

def send_vendor_inquiry(
    vendor_emails: List[str],
    lead_no: str,
    rfq_ref: str,
    project: str,
    attachment_paths: List[str],
    smtp_host: str = "",
    smtp_port: int = 587,
    smtp_user: str = "",
    smtp_pass: str = "",
    from_addr: str = "",
) -> dict:
    """
    Phase 3.4: Send the vendor inquiry package to a list of vendor emails.

    Returns {"sent": [...], "failed": [...]} for each address.

    If SMTP credentials are not configured, returns a dry-run result so the
    pipeline does not crash in development.
    """
    if not smtp_host or not smtp_user:
        log.warning(
            "SMTP not configured — skipping vendor email send "
            "(SMTP_HOST / SMTP_USER env vars not set)."
        )
        return {"sent": [], "failed": vendor_emails, "dry_run": True}

    sent, failed = [], []
    subject = (
        f"Vendor Inquiry — Cenlub Systems | Lead: {lead_no} | {rfq_ref} / {project}"
    )
    body = (
        f"Dear Vendor,\n\n"
        f"Please find attached the Vendor Inquiry for Lead No. {lead_no}.\n"
        f"Project: {project}  |  RFQ Ref: {rfq_ref}\n\n"
        f"Kindly fill in the Make, Unit Price, Delivery Time, Offer Reference, "
        f"and Validity Date columns in the attached Excel sheet and return "
        f"it to us at your earliest convenience.\n\n"
        f"Regards,\nCenlub Systems — Engineering Team"
    )

    for addr in vendor_emails:
        try:
            msg = MIMEMultipart()
            msg["From"] = from_addr or smtp_user
            msg["To"] = addr
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain"))

            for path in attachment_paths:
                if os.path.exists(path):
                    with open(path, "rb") as fh:
                        part = MIMEApplication(fh.read(), Name=os.path.basename(path))
                    part["Content-Disposition"] = (
                        f'attachment; filename="{os.path.basename(path)}"'
                    )
                    msg.attach(part)

            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo()
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, addr, msg.as_string())
            sent.append(addr)
            log.info("Vendor inquiry sent to %s", addr)
        except Exception as exc:
            log.error("Failed to send to %s: %s", addr, exc)
            failed.append(addr)

    return {"sent": sent, "failed": failed, "dry_run": False}


# ── Auto-reminder (Phase 3.5) ─────────────────────────────────────────────────

def send_vendor_reminder(
    vendor_emails: List[str],
    lead_no: str,
    rfq_ref: str,
    project: str,
    sent_date: date,
    sla_days: int = 3,
    smtp_host: str = "",
    smtp_port: int = 587,
    smtp_user: str = "",
    smtp_pass: str = "",
    from_addr: str = "",
) -> dict:
    """
    Phase 3.5: Send a reminder if the vendor has not responded within *sla_days*.
    Returns same {"sent", "failed", "dry_run"} dict as send_vendor_inquiry().
    """
    today = date.today()
    due_date = sent_date + timedelta(days=sla_days)
    if today < due_date:
        log.info(
            "SLA not yet breached for %s (due %s). No reminder sent.",
            lead_no, due_date,
        )
        return {"sent": [], "failed": [], "dry_run": False, "sla_due": str(due_date)}

    if not smtp_host or not smtp_user:
        log.warning("SMTP not configured — skipping vendor reminder.")
        return {"sent": [], "failed": vendor_emails, "dry_run": True}

    sent, failed = [], []
    subject = (
        f"REMINDER — Vendor Inquiry | Lead: {lead_no} | {rfq_ref} / {project}"
    )
    body = (
        f"Dear Vendor,\n\n"
        f"This is a reminder regarding our vendor inquiry for Lead No. {lead_no}.\n"
        f"Project: {project}  |  RFQ Ref: {rfq_ref}\n\n"
        f"We have not yet received your quotation. Kindly revert at your earliest "
        f"with pricing, delivery time, and offer validity.\n\n"
        f"Regards,\nCenlub Systems — Engineering Team"
    )

    for addr in vendor_emails:
        try:
            msg = MIMEMultipart()
            msg["From"] = from_addr or smtp_user
            msg["To"] = addr
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain"))

            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo()
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, addr, msg.as_string())
            sent.append(addr)
            log.info("Reminder sent to %s for lead %s", addr, lead_no)
        except Exception as exc:
            log.error("Reminder failed for %s: %s", addr, exc)
            failed.append(addr)

    return {"sent": sent, "failed": failed, "dry_run": False}
