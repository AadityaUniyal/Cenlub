"""
tests/test_pipeline.py
======================
Integration tests for the unified pipeline.
Run:  python -m pytest tests/ -v
      python tests/test_pipeline.py
"""

import logging
import os
import sys

# Ensure project root is on the path
sys.path.insert(
    0, os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
)

from src.core.models import (  # noqa: E402
    MandatorySpare,
    QuadraticHashTable,
    RMCData,
)
from src.core.validators import (  # noqa: E402
    confidence_score,
    sanity_check,
)
from src.extraction.bom_extractor import (  # noqa: E402
    clean_price,
    split_tags,
)
from src.extraction.classifier import classify  # noqa: E402
from src.extraction.rmc_extractor import (  # noqa: E402
    parse_spares_table,
    split_email_thread,
)
from src.ingestion.loader import load_pdf  # noqa: E402
from src.output.excel_writer import write_excel  # noqa: E402

logging.basicConfig(level=logging.INFO)
RAW = "data/raw"
OUT = "data/processed_test"
os.makedirs(OUT, exist_ok=True)


# ── Unit tests ────────────────────────────────────────────────────────────


def test_quadratic_hash_table():
    ht = QuadraticHashTable()
    assert ht.add("alpha") is True      # new
    assert ht.add("alpha") is False     # duplicate
    assert ht.add("beta") is True
    assert ht.contains("alpha") is True
    assert ht.contains("gamma") is False
    # Stress: 200 unique keys
    for i in range(200):
        ht.add(f"key_{i}")
    for i in range(200):
        assert ht.contains(f"key_{i}"), (
            f"key_{i} missing after bulk insert"
        )
    print("  ✓ QuadraticHashTable")


def test_split_tags():
    assert split_tags("P-101A/B") == ["P-101A", "P-101B"]
    assert split_tags("XV-101/102/103") == ["XV-101", "XV-102", "XV-103"]
    assert split_tags("XV-101A/B/C") == ["XV-101A", "XV-101B", "XV-101C"]
    assert split_tags("XV-101") == ["XV-101"]
    assert split_tags("") == []
    print("  ✓ split_tags")


def test_clean_price():
    assert clean_price("otIoNrRs 75,000") == "INR 75,000"
    assert clean_price("$ 1,20,000") == "USD 1,20,000"
    assert clean_price("INR 35,000") == "INR 35,000"
    assert clean_price("100") == "INR 100"
    print("  ✓ clean_price")


def test_email_thread_split():
    sample = (
        "From: buyer@co.com\nPlease update flow to 60 LPM.\n\n"
        "-----Original Message-----\n"
        "From: buyer@co.com\nInitial request: flow 50 LPM."
    )
    segs = split_email_thread(sample)
    assert len(segs) == 2
    assert "60 LPM" in segs[0]
    assert "50 LPM" in segs[1]
    print("  ✓ split_email_thread")


def test_moc_normalisation():
    from src.core.models import standardise_moc
    assert standardise_moc("304 stainless") == "SS304"
    assert standardise_moc("ss-316") == "SS316"
    assert standardise_moc("mild steel") == "Carbon Steel"
    assert standardise_moc("copper tubes") == "Copper"
    assert standardise_moc("XX") == "XX"
    print("  ✓ MOC normalisation")


def test_rmc_field_validators():
    rmc = RMCData(
        flow_rate="50 lpm",
        pressure="10 bar",
        temperature="45 deg c",
        filtration_level="10 microns",
    )
    assert rmc.flow_rate == "50 LPM"
    assert rmc.pressure == "10 Bar"
    assert rmc.temperature == "45 Deg C"
    assert rmc.filtration_level == "10 Microns"
    print("  ✓ RMCData field validators")


def test_sanity_checks_pass():
    rmc = RMCData(
        flow_rate="50 LPM", pressure="10 Bar",
        temperature="45 Deg C", filtration_level="10 Microns",
    )
    assert sanity_check(rmc) == []
    print("  ✓ sanity_check — all pass")


def test_sanity_checks_fail():
    rmc = RMCData(
        flow_rate="4500 LPM", pressure="250 Bar",
        temperature="450 Deg C", filtration_level="250 Microns",
    )
    warns = sanity_check(rmc)
    assert len(warns) == 4
    print("  ✓ sanity_check — 4 boundary violations detected")


def test_confidence_score():
    full_rmc = RMCData(
        lead_no="P001", project_name="TEST", customer="ACME",
        application="Lube Skid", industry="Power",
        flow_rate="50 LPM", pressure="10 Bar",
        temperature="45 Deg C", oil_grade="ISO VG 46",
        filtration_level="10 Microns", heat_exchanger_moc="Copper",
        reservoir_moc="SS304", control_panel_requirement="PLC",
    )
    assert confidence_score(full_rmc) == 100.0
    empty_rmc = RMCData()
    assert confidence_score(empty_rmc) == 0.0
    print("  ✓ confidence_score")


def test_parse_spares_table():
    table = [
        ["Item No", "Description", "Qty", "Part Number"],
        ["1", "Rotor Assembly", "2", "ROT-101"],
        ["2", "Mechanical Seal", "4", "MS-202"],
    ]
    spares = parse_spares_table(table, "mandatory")
    assert len(spares) == 2
    assert spares[0]["description"] == "Rotor Assembly"
    assert spares[0]["part_no"] == "ROT-101"
    print("  ✓ parse_spares_table")


# ── Integration tests ─────────────────────────────────────────────────────


def test_ingestion_and_classification():
    mock_files = [
        "rfq_email.pdf", "los_specification.pdf", "technical_spec.pdf",
        "bom_sheet.pdf", "mandatory_spares.pdf", "om_spares.pdf",
    ]
    expected = {
        "rfq_email.pdf":         "Email",
        "los_specification.pdf": "LOS Specification",
        "technical_spec.pdf":    "Technical Specification",
        "bom_sheet.pdf":         "BOM",
        "mandatory_spares.pdf":  "Mandatory Spares",
        "om_spares.pdf":         "O&M Spares",
    }
    for fname in mock_files:
        path = os.path.join(RAW, fname)
        assert os.path.exists(path), (
            f"Missing mock file: {path}. Run generate_mock_tenders.py."
        )
        doc = load_pdf(path)
        assert len(doc["full_text"]) > 50
        cls = classify(doc["full_text"])
        assert cls == expected[fname], (
            f"{fname}: expected {expected[fname]}, got {cls}"
        )
    print("  ✓ ingestion + classification")


def test_full_pipeline():
    """End-to-end pipeline producing a valid CSV."""
    from pipeline import run_pipeline
    result = run_pipeline(
        input_path=RAW,
        output_csv=os.path.join(OUT, "master_output_test.csv"),
    )
    rmc = result["rmc"]
    val = result["validation"]

    assert rmc.lead_no == "P26060840"
    assert rmc.project_name == "FLOWMORE"
    assert rmc.customer == "TATA POWER"
    assert rmc.flow_rate == "50 LPM"
    assert rmc.pressure == "10 Bar"
    assert rmc.temperature == "45 Deg C"
    assert rmc.oil_grade == "ISO VG 46"
    assert rmc.filtration_level == "10 Microns"
    assert rmc.reservoir_moc == "SS304"
    assert len(rmc.mandatory_spares) == 3
    assert len(rmc.om_spares) == 4

    assert val["confidence_score"] == 100.0
    assert os.path.exists(result["output_csv"])

    with open(result["output_csv"], encoding="utf-8") as f:
        content = f.read()
    assert "SECTION A" in content
    assert "SECTION B" in content
    assert "SECTION C" in content
    assert "SECTION D" in content
    assert "FLOWMORE" in content
    assert "Rotor Assembly" in content
    assert "Screw Pump" in content
    print("  ✓ full end-to-end pipeline")


def test_excel_output():
    import openpyxl
    rmc = RMCData(
        lead_no="TEST", project_name="TESTPROJ",
        mandatory_spares=[
            MandatorySpare(
                description="Test Part", qty="1", part_no="TP-001"
            )
        ],
    )
    out = os.path.join(OUT, "test_rmc.xlsx")
    write_excel(rmc, out)
    wb = openpyxl.load_workbook(out)
    assert "Master RMC" in wb.sheetnames
    assert "Mandatory Spares" in wb.sheetnames
    assert "O&M Spares" in wb.sheetnames
    print("  ✓ Excel output")


# ── Runner ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n=== Running Pipeline Tests ===\n")
    test_quadratic_hash_table()
    test_split_tags()
    test_clean_price()
    test_email_thread_split()
    test_moc_normalisation()
    test_rmc_field_validators()
    test_sanity_checks_pass()
    test_sanity_checks_fail()
    test_confidence_score()
    test_parse_spares_table()
    test_ingestion_and_classification()
    test_full_pipeline()
    test_excel_output()
    print("\n=== ALL TESTS PASSED ===\n")
