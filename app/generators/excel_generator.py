import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.schema.models import RMCData


def style_sheet_headers(ws, title_text, fill_color="1F4E78"):
    """Applies standard title banner styling to the top of a worksheet."""
    # Create Title Block
    ws.merge_cells("A1:D2")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(
        start_color=fill_color,
        end_color=fill_color,
        fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20


def apply_borders_and_font(ws, font_name="Calibri", font_size=11):
    """Applies a clean grid border and consistent font styling to all populated cells."""
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.font = Font(name=font_name, size=font_size)
                cell.border = thin_border


def autofit_column_widths(ws):
    """Dynamically scales column widths to match contents."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Avoid measuring merged title cell width to prevent extreme
            # columns
            if cell.row in [1, 2]:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def generate_excel_output(data: RMCData, output_path: str):
    """
    Creates a styled Excel workbook from RMCData.
    Tabs: Master RMC, Mandatory Spares, O&M Spares.
    """
    wb = openpyxl.Workbook()

    # 1. Master RMC Sheet
    ws1 = wb.active
    ws1.title = "Master RMC"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner (Steel Blue)
    style_sheet_headers(
        ws1, f"Requirement Mapping Checklist (RMC) - Lead: {data.lead_no}", "1F4E78")

    # Parameter Headers
    ws1["A4"] = "Engineering Parameter"
    ws1["B4"] = "Extracted Value"
    for col in ["A4", "B4"]:
        cell = ws1[col]
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="2F5597",
            end_color="2F5597",
            fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Add parameter data
    parameters = [
        ("Lead Number", data.lead_no),
        ("Project Name", data.project_name),
        ("Customer Name", data.customer),
        ("Application", data.application),
        ("Industry", data.industry),
        ("Flow Rate", data.flow_rate),
        ("Oil Grade", data.oil_grade),
        ("Operating Pressure", data.pressure),
        ("Operating Temperature", data.temperature),
        ("Filtration Level", data.filtration_level),
        ("Heat Exchanger MOC", data.heat_exchanger_moc),
        ("Reservoir MOC", data.reservoir_moc),
        ("Control Panel Requirement", data.control_panel_requirement),
    ]

    for idx, (param, val) in enumerate(parameters):
        row_idx = 5 + idx
        ws1.cell(row=row_idx, column=1, value=param)
        val_cell = ws1.cell(row=row_idx, column=2, value=val)
        # Highlight values that are placeholder XX
        if val == "XX":
            val_cell.fill = PatternFill(
                start_color="FFE6E6",
                end_color="FFE6E6",
                fill_type="solid")
            val_cell.font = Font(color="C00000")

    apply_borders_and_font(ws1)
    autofit_column_widths(ws1)

    # 2. Mandatory Spares Tab
    ws2 = wb.create_sheet(title="Mandatory Spares")
    ws2.views.sheetView[0].showGridLines = True
    style_sheet_headers(
        ws2,
        "Mandatory Spares Checklist",
        "C53030")  # Dark Red

    # Spares Table headers
    spares_headers = ["Item No", "Description", "Quantity", "Part Number"]
    for col_idx, header in enumerate(spares_headers):
        cell = ws2.cell(row=4, column=col_idx + 1, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="9B2C2C",
            end_color="9B2C2C",
            fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r_idx, spare in enumerate(data.mandatory_spares):
        row_num = 5 + r_idx
        ws2.cell(row=row_num, column=1, value=spare.item_no)
        ws2.cell(row=row_num, column=2, value=spare.description)
        ws2.cell(row=row_num, column=3, value=spare.qty)
        ws2.cell(row=row_num, column=4, value=spare.part_no)

    apply_borders_and_font(ws2)
    autofit_column_widths(ws2)

    # 3. O&M Spares Tab
    ws3 = wb.create_sheet(title="O&M Spares")
    ws3.views.sheetView[0].showGridLines = True
    style_sheet_headers(
        ws3,
        "Operation & Maintenance (O&M) Spares Checklist",
        "D69E2E")  # Gold/Amber

    om_headers = [
        "Item No",
        "Description",
        "Quantity",
        "Recommended Frequency"]
    for col_idx, header in enumerate(om_headers):
        cell = ws3.cell(row=4, column=col_idx + 1, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="B7791F",
            end_color="B7791F",
            fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r_idx, spare in enumerate(data.om_spares):
        row_num = 5 + r_idx
        ws3.cell(row=row_num, column=1, value=spare.item_no)
        ws3.cell(row=row_num, column=2, value=spare.description)
        ws3.cell(row=row_num, column=3, value=spare.qty)
        ws3.cell(row=row_num, column=4, value=spare.frequency)

    apply_borders_and_font(ws3)
    autofit_column_widths(ws3)

    # Save Workbook
    wb.save(output_path)
